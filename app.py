if __name__ == "__main__":
    import eventlet

    eventlet.monkey_patch()

import os
from dotenv import load_dotenv

# Load variables from a .env file if present for local development. In
# production (e.g. Fly.io) the environment is configured through secrets.
load_dotenv()

if os.environ.get("FLASK_RUN_FROM_CLI") and not os.environ.get("ALLOW_FLASK_CLI"):
    raise RuntimeError(
        "❌ Не запускай через 'flask run'. Используй только 'python app.py'."
    )

import csv
import json
import re
import threading
import time
import uuid

from flask_socketio import SocketIO
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO

import openpyxl
import pandas as pd
from flask import (
    Flask,
    flash,
    jsonify,
    abort,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_login import (
    LoginManager,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_migrate import Migrate
from sqlalchemy import text, inspect
from shapely.geometry import shape
from types import SimpleNamespace
from werkzeug.security import check_password_hash, generate_password_hash

from config import Config
from geocode import geocode_address
from models import (
    Courier,
    DeliveryZone,
    ImportJob,
    Order,
    User,
    ImportBatch,
    WorkArea,
    db,
)


app = Flask(__name__)
app.config.from_object("config.Config")

# Read the connection string from environment variables. Fly.io provides
# secrets as environment variables, so just looking them up is enough.
database_uri = os.getenv("SQLALCHEMY_DATABASE_URI") or os.getenv("DATABASE_URL")

if not database_uri:
    raise RuntimeError("Database connection string is not set")

# SQLAlchemy 2.x requires an explicit driver name. Convert legacy URLs if necessary so that `postgres://` continues to work
if database_uri.startswith("postgres://"):
    database_uri = database_uri.replace("postgres://", "postgresql+psycopg2://", 1)
elif database_uri.startswith("postgresql://") and not database_uri.startswith(
    "postgresql+psycopg2://"
):
    database_uri = database_uri.replace("postgresql://", "postgresql+psycopg2://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = database_uri
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db.init_app(app)
migrate = Migrate(app, db)

socketio = SocketIO(app, async_mode="eventlet")

# store row-level import errors for each job
job_errors = defaultdict(list)

login_manager = LoginManager(app)
login_manager.login_view = "login"


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


def admin_required(func):
    from functools import wraps

    @wraps(func)
    @login_required
    def wrapper(*args, **kwargs):
        if current_user.role != "admin":
            return redirect(url_for("orders"))
        return func(*args, **kwargs)

    return wrapper


def point_in_polygon(x, y, polygon):
    inside = False
    n = len(polygon)
    p1x, p1y = polygon[0]
    for i in range(1, n + 1):
        p2x, p2y = polygon[i % n]
        if y > min(p1y, p2y):
            if y <= max(p1y, p2y):
                if x <= max(p1x, p2x):
                    if p1y != p2y:
                        xinters = (y - p1y) * (p2x - p1x) / (p2y - p1y) + p1x
                    if p1x == p2x or x <= xinters:
                        inside = not inside
        p1x, p1y = p2x, p2y
    return inside


def detect_zone(lat, lng):
    zones = DeliveryZone.query.all()
    for zone in zones:
        try:
            polygon = json.loads(zone.polygon_json) if zone.polygon_json else []
        except Exception:
            polygon = []
        if polygon and point_in_polygon(lng, lat, polygon):
            return zone.name
    return None


def assign_courier_for_zone(zone_name):
    """Return Courier instance serving the given zone name."""
    if not zone_name:
        return None
    couriers = Courier.query.all()
    for c in couriers:
        if c.zones:
            zones = [z.strip() for z in c.zones.split(",") if z.strip()]
            if zone_name in zones:
                return c
    return None


def populate_demo_data():
    """Populate database with demo zones, couriers and orders."""
    if Order.query.first():
        return

    zones = []
    for i in range(5):
        polygon = [
            [37.6 + i * 0.05, 55.7],
            [37.65 + i * 0.05, 55.7],
            [37.65 + i * 0.05, 55.75],
            [37.6 + i * 0.05, 55.75],
        ]
        zone = DeliveryZone(
            name=f"Zone {i + 1}",
            color=f"#33{i}{i}ff" if i < 10 else "#3388ff",
            polygon_json=json.dumps(polygon),
        )
        zones.append(zone)
    db.session.add_all(zones)
    db.session.commit()

    couriers_to_add = []
    users = []

    if not User.query.filter_by(username="admin").first():
        users.append(
            User(
                username="admin",
                password_hash=generate_password_hash("admin"),
                role="admin",
            )
        )

    for i in range(5):
        courier_name = f"Courier {i+1}"
        courier = Courier.query.filter_by(name=courier_name).first()
        if not courier:
            courier = Courier(
                name=courier_name,
                telegram=f"@courier{i+1}",
                zones=f"Zone {i+1}",
            )
            couriers_to_add.append(courier)

        if not User.query.filter_by(username=f"courier{i+1}").first():
            users.append(
                User(
                    username=f"courier{i+1}",
                    password_hash=generate_password_hash("courier"),
                    role="courier",
                )
            )

    if couriers_to_add:
        db.session.add_all(couriers_to_add)
    if users:
        db.session.add_all(users)
    db.session.commit()

    couriers = [
        Courier.query.filter_by(name=f"Courier {i+1}").first() for i in range(5)
    ]

    for i in range(20):
        zone = zones[i % 5]
        lat = 55.71 + (i % 5) * 0.01
        lng = 37.61 + (i % 5) * 0.01
        order = Order(
            order_number=f"ORD{100 + i}",
            client_name=f"Client {i+1}",
            phone=f"+700000000{(i+1):02d}",
            address=f"Demo address {i+1}",
            latitude=lat,
            longitude=lng,
            zone=zone.name,
            import_batch="DEMO",
        )
        db.session.add(order)
        order.courier = couriers[i % 5]
    db.session.commit()


with app.app_context():
    db.create_all()
    if db.engine.url.drivername.startswith("sqlite"):
        with db.engine.connect() as conn:
            table_info = conn.execute(text("PRAGMA table_info(orders)")).fetchall()
            column_names = [col[1] for col in table_info]

            if "problem_comment" not in column_names:
                conn.execute(text("ALTER TABLE orders ADD COLUMN problem_comment TEXT"))
    # populate_demo_data()  # Demo data generation disabled


@app.route("/")
def index():
    if current_user.is_authenticated:
        if current_user.role == "courier":
            return redirect(url_for("courier_dashboard"))
        return redirect(url_for("orders"))
    else:
        return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            flash("Вы вошли в систему", "success")
            if user.role == "courier":
                return redirect(url_for("courier_dashboard"))
            return redirect(url_for("orders"))
        flash("Неверные логин или пароль", "warning")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Вы вышли из системы", "success")
    return redirect(url_for("login"))


@app.route("/orders", methods=["GET", "POST"])
@login_required
def orders():
    if current_user.role == "courier":
        return redirect(url_for("courier_dashboard"))
    if request.method == "POST":
        order_id = request.form["id"]
        status = request.form["status"]
        order = Order.query.get(order_id)
        if order:
            order.status = status
            if status == "Доставлен":
                order.delivered_at = date.today()
            else:
                order.delivered_at = None
            db.session.commit()
            flash("Статус заказа обновлён", "success")
        return redirect(url_for("orders"))
    courier = None
    allowed_zones = []
    if current_user.role == "courier":
        courier = Courier.query.filter_by(telegram=f"@{current_user.username}").first()
        if courier and courier.zones:
            allowed_zones = [z.strip() for z in courier.zones.split(",") if z.strip()]

    couriers_list = Courier.query.all()
    couriers_map = {c.id: c for c in couriers_list}

    orders_by_batch = {}
    orders_by_zone = defaultdict(list)
    all_orders = []

    batches = ImportBatch.query.order_by(ImportBatch.created_at.desc()).all()
    for batch in batches:
        query = Order.query.filter_by(import_batch_id=batch.id)
        if allowed_zones:
            query = query.filter(Order.zone.in_(allowed_zones))
        elif current_user.role == "courier":
            query = query.filter(db.text("0=1"))
        items = query.order_by(Order.id).all()
        orders_by_batch[batch] = items
        for o in items:
            key = getattr(o, "zone", None) or "Не определена"
            orders_by_zone[key].append(o)
        all_orders.extend(items)

    zones_list = DeliveryZone.query.all()
    zones_dict = [
        {
            "id": z.id,
            "name": z.name,
            "color": z.color,
            "polygon": json.loads(z.polygon_json),
        }
        for z in zones_list
    ]

    return render_template(
        "orders.html",
        orders=all_orders,
        orders_by_zone=orders_by_zone,
        orders_by_batch=orders_by_batch,
        couriers=couriers_list,
        zones=zones_dict,
    )


@app.route("/orders/<int:order_id>/update", methods=["POST"])
@login_required
def update_order(order_id):
    order = Order.query.get_or_404(order_id)
    old_address = order.address
    order.client_name = request.form.get("client_name", order.client_name)
    order.phone = request.form.get("phone", order.phone)
    order.address = request.form.get("address", order.address)
    order.note = request.form.get("note", order.note)
    lat_val = request.form.get("latitude")
    lon_val = request.form.get("longitude")
    manual_coords = False
    if lat_val and lon_val:
        try:
            lat = float(lat_val)
            lon = float(lon_val)
            order.latitude = lat
            order.longitude = lon
            manual_coords = True
        except ValueError:
            flash("Некорректные координаты", "warning")
    if order.address != old_address and not manual_coords:
        lat, lng = geocode_address(order.address)
        order.latitude = lat
        order.longitude = lng
        if lat and lng:
            order.zone = detect_zone(lat, lng)
        else:
            order.zone = None
            flash("Не удалось определить координаты по адресу", "warning")
    if manual_coords and order.latitude and order.longitude:
        order.zone = detect_zone(order.latitude, order.longitude)
    courier_val = request.form.get("courier_id")
    if courier_val:
        try:
            order.courier_id = int(courier_val)
        except ValueError:
            order.courier_id = None
    elif order.zone:
        c = assign_courier_for_zone(order.zone)
        db.session.add(order)
        order.courier = c
    db.session.commit()
    flash("Заказ обновлен", "success")
    return redirect(url_for("orders"))


@app.route("/orders/<int:order_id>/set_coords", methods=["GET", "POST"])
@login_required
def set_coords(order_id):
    order = Order.query.get_or_404(order_id)
    if request.method == "POST":
        lat = request.form.get("latitude")
        lng = request.form.get("longitude")
        try:
            lat = float(lat)
            lng = float(lng)
        except (TypeError, ValueError):
            flash("Некорректные координаты", "warning")
            return redirect(url_for("set_coords", order_id=order_id))
        order.latitude = lat
        order.longitude = lng
        order.zone = detect_zone(lat, lng)
        c = assign_courier_for_zone(order.zone)
        db.session.add(order)
        order.courier = c
        db.session.commit()
        flash("Координаты сохранены", "success")
        return redirect(url_for("orders"))
    zones_list = DeliveryZone.query.all()
    zones_dict = [
        {
            "id": z.id,
            "name": z.name,
            "color": z.color,
            "polygon": json.loads(z.polygon_json),
        }
        for z in zones_list
    ]
    return render_template("set_coords.html", order=order, zones=zones_dict)


@app.route("/orders/set_point", methods=["POST"])
@login_required
def set_point():
    order_id = request.form.get("order_id")
    lat = request.form.get("lat")
    lon = request.form.get("lon")
    try:
        order_id = int(order_id)
        lat = float(lat)
        lon = float(lon)
    except (TypeError, ValueError):
        return jsonify({"success": False}), 400
    order = Order.query.get_or_404(order_id)
    order.latitude = lat
    order.longitude = lon
    order.zone = detect_zone(lat, lon)
    c = assign_courier_for_zone(order.zone)
    db.session.add(order)
    order.courier = c
    db.session.commit()
    return jsonify(
        {
            "success": True,
            "zone": order.zone,
            "courier": order.courier.name if order.courier else None,
        }
    )


@app.route("/api/orders/<int:order_id>/coordinates", methods=["POST"])
@login_required
def api_set_coordinates(order_id):
    data = request.get_json(silent=True) or {}
    lat = data.get("latitude")
    lon = data.get("longitude")
    try:
        lat = float(lat)
        lon = float(lon)
    except (TypeError, ValueError):
        return jsonify({"success": False}), 400
    order = Order.query.get_or_404(order_id)
    order.latitude = lat
    order.longitude = lon
    order.zone = detect_zone(lat, lon)
    courier = assign_courier_for_zone(order.zone)
    db.session.add(order)
    order.courier = courier
    db.session.commit()
    return jsonify(
        {
            "success": True,
            "zone": order.zone,
            "courier": order.courier.name if order.courier else None,
        }
    )


@app.route("/orders/set_coordinates/<int:order_id>", methods=["POST"])
@login_required
def set_coordinates(order_id):
    return api_set_coordinates(order_id)


@login_required
def add_comment_photo(order_id):
    if current_user.role != "courier":
        return redirect(url_for("orders"))
    order = Order.query.get_or_404(order_id)
    comment = request.form.get("comment", "").strip()
    file = request.files.get("photo")
    if file and file.filename:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in [".jpg", ".jpeg", ".png"]:
            flash("Допустимы только файлы JPG и PNG", "warning")
            return redirect(url_for("orders"))
        upload_dir = os.path.join(app.static_folder, "uploads")
        os.makedirs(upload_dir, exist_ok=True)
        filename = f"order_{order_id}_{int(time.time())}.jpg"
        path = os.path.join(upload_dir, filename)
        file.save(path)
        order.photo_filename = filename
    if comment:
        order.comment = comment
    db.session.commit()
    flash("Комментарий сохранен", "success")
    return redirect(url_for("orders"))


@app.route("/orders/<int:order_id>/delete", methods=["POST"])
@admin_required
def delete_order(order_id):
    """Delete a single order."""
    order = Order.query.get_or_404(order_id)
    db.session.delete(order)
    db.session.commit()
    flash("Заказ удалён", "success")
    return redirect(url_for("orders"))


@app.route("/delete_batch/<int:batch_id>", methods=["POST"])
@admin_required
def delete_batch(batch_id):
    """Delete all orders imported in the given batch."""
    batch = ImportBatch.query.get_or_404(batch_id)
    Order.query.filter_by(import_batch_id=batch.id).delete()
    db.session.delete(batch)
    db.session.commit()
    flash("Таблица удалена", "success")
    return redirect(url_for("orders"))


@app.route("/map")
@login_required
def map_view():
    query = Order.query
    if current_user.role == "courier":
        courier = Courier.query.filter_by(telegram=f"@{current_user.username}").first()
        zones = []
        if courier and courier.zones:
            zones = [z.strip() for z in courier.zones.split(",") if z.strip()]
        if zones:
            query = query.filter(Order.zone.in_(zones))
        else:
            query = query.filter(db.text("0=1"))
    orders = query.all()
    orders_dict = [
        {
            "id": o.id,
            "address": o.address,
            "lat": o.latitude,
            "lng": o.longitude,
            "zone": o.zone,
            "status": o.status,
        }
        for o in orders
    ]

    work_area = WorkArea.query.first()
    wa_json = None
    if work_area:
        try:
            wa_dict = json.loads(work_area.geojson)
            wa_json = wa_dict.get("geometry")
        except Exception:
            wa_json = None

    zones = DeliveryZone.query.all()
    zones_dict = []
    for z in zones:
        try:
            poly = json.loads(z.polygon_json) if z.polygon_json else []
        except Exception:
            poly = []
        zones_dict.append({
            "id": z.id,
            "name": z.name,
            "color": z.color,
            "polygon": poly,
        })

    return render_template(
        "map.html",
        orders=orders_dict,
        zones=zones_dict,
        workarea=wa_json,
    )


@app.route("/workarea", methods=["GET", "POST"])
@admin_required
def workarea():
    area = WorkArea.query.first()
    if request.method == "POST":
        color = request.form.get("color") or "#777777"
        geojson = request.form.get("geojson") or "{}"
        if area:
            area.color = color
            area.geojson = geojson
        else:
            area = WorkArea(name="Рабочая область", color=color, geojson=geojson)
            db.session.add(area)
        db.session.commit()
        flash("Рабочая область сохранена", "success")
        return redirect(url_for("workarea"))
    if not area:
        area = WorkArea(
            name="Рабочая область",
            color="#777777",
            geojson=json.dumps(
                {"type": "Feature", "geometry": {"type": "Polygon", "coordinates": []}}
            ),
        )
    return render_template("workarea/index.html", area=area)


@app.route("/zones")
@admin_required
def zones():
    work_area = WorkArea.query.first()
    wa_json = None
    if work_area:
        try:
            wa_dict = json.loads(work_area.geojson)
            wa_json = wa_dict.get("geometry")
        except Exception:
            wa_json = None

    zones = DeliveryZone.query.all()
    zones_dict = []
    for z in zones:
        try:
            poly = json.loads(z.polygon_json) if z.polygon_json else []
        except Exception:
            poly = []
        zones_dict.append({
            "id": z.id,
            "name": z.name,
            "color": z.color,
            "polygon": poly,
        })

    return render_template("zones.html", zones=zones_dict, workarea=wa_json)


@app.route("/zones/new", methods=["GET", "POST"])
@admin_required
def new_zone():
    if request.method == "POST":
        name = request.form["name"]
        color = request.form["color"]
        polygon = request.form["geojson"]

        coords = json.loads(polygon) if polygon else []
        zone_geo = {
            "type": "Feature",
            "geometry": {"type": "Polygon", "coordinates": [coords]},
        }
        work_area = WorkArea.query.first()
        wa_json = None
        if work_area:
            try:
                wa_dict = json.loads(work_area.geojson)
                wa_json = wa_dict.get("geometry")
            except Exception:
                wa_json = None
        if work_area and wa_json and not shape(wa_json).contains(
            shape(zone_geo["geometry"])
        ):
            flash("Зона должна быть внутри рабочей области", "danger")
            return redirect(url_for("zones"))

        zone = DeliveryZone(name=name, color=color, polygon_json=polygon)
        db.session.add(zone)
        db.session.commit()
        flash("Зона создана", "success")
        return redirect(url_for("zones"))

    zone = DeliveryZone(name="", color="#3388ff", polygon_json="[]")
    work_area = WorkArea.query.first()
    wa_json = None
    work_color = "#777777"
    if work_area:
        try:
            wa_dict = json.loads(work_area.geojson)
            wa_json = wa_dict.get("geometry")
        except Exception:
            wa_json = None
        work_color = work_area.color
    zones = DeliveryZone.query.all()
    zones_dict = []
    for z in zones:
        try:
            poly = json.loads(z.polygon_json) if z.polygon_json else []
        except Exception:
            poly = []
        zones_dict.append({
            "id": z.id,
            "name": z.name,
            "color": z.color,
            "polygon": poly,
        })
    return render_template(
        "edit_zone.html",
        zone=zone,
        new=True,
        zones=zones_dict,
        workarea=wa_json,
        workcolor=work_color,
    )


@app.route("/zones/<int:zone_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_zone(zone_id):
    zone = DeliveryZone.query.get_or_404(zone_id)
    if request.method == "POST":
        zone.name = request.form.get("name", zone.name)
        zone.color = request.form.get("color", zone.color)
        polygon = request.form.get("polygon")
        if polygon is not None:
            coords = json.loads(polygon) if polygon else []
            zone_geo = {
                "type": "Feature",
                "geometry": {"type": "Polygon", "coordinates": [coords]},
            }
            work_area = WorkArea.query.first()
            wa_json = None
            if work_area:
                try:
                    wa_dict = json.loads(work_area.geojson)
                    wa_json = wa_dict.get("geometry")
                except Exception:
                    wa_json = None
            if work_area and wa_json and not shape(wa_json).contains(
                shape(zone_geo["geometry"])
            ):
                flash("Зона должна быть внутри рабочей области", "danger")
                return redirect(url_for("zones"))
            zone.polygon_json = polygon
        db.session.commit()
        flash("Зона обновлена", "success")
        return redirect(url_for("zones"))
    work_area = WorkArea.query.first()
    wa_json = None
    work_color = "#777777"
    if work_area:
        try:
            wa_dict = json.loads(work_area.geojson)
            wa_json = wa_dict.get("geometry")
        except Exception:
            wa_json = None
        work_color = work_area.color
    return render_template(
        "edit_zone.html",
        zone=zone,
        workarea=wa_json,
        workcolor=work_color,
    )


@app.route("/zones/<int:zone_id>/delete")
@admin_required
def delete_zone(zone_id):
    zone = DeliveryZone.query.get_or_404(zone_id)
    Order.query.filter_by(zone=zone.name).update({"zone": None, "courier_id": None})
    db.session.delete(zone)
    db.session.commit()
    flash("Зона удалена", "success")
    return redirect(url_for("zones"))


@app.route("/couriers")
@admin_required
def couriers():
    couriers = Courier.query.all()
    return render_template("couriers.html", couriers=couriers)


@app.route("/courier")
@login_required
def courier_dashboard():
    if current_user.role != "courier":
        return redirect(url_for("orders"))
    courier = Courier.query.filter_by(telegram=f"@{current_user.username}").first()
    if not courier:
        abort(403)
    orders = (
        Order.query.filter_by(courier_id=courier.id)
        .filter(
            Order.status.in_(
                ["Подготовлен к доставке", "Выдано на доставку", "Проблема"]
            )
        )
        .order_by(Order.id)
        .all()
    )
    prepared_count = sum(1 for o in orders if o.status == "Подготовлен к доставке")
    delivered_orders = (
        Order.query.filter_by(courier_id=courier.id, status="Доставлен")
        .order_by(Order.delivered_at.desc())
        .all()
    )
    all_orders = [
        {
            "id": o.id,
            "order_number": o.order_number,
            "address": o.address,
            "lat": o.latitude,
            "lng": o.longitude,
            "status": o.status,
        }
        for o in Order.query.filter_by(courier_id=courier.id)
        .filter(
            Order.status.in_(
                ["Подготовлен к доставке", "Выдано на доставку", "Проблема"]
            )
        )
        .all()
    ]
    return render_template(
        "courier.html",
        orders=orders,
        prepared_count=prepared_count,
        delivered_orders=delivered_orders,
        all_orders=all_orders,
    )


@app.route("/courier/take", methods=["POST"])
@login_required
def courier_take():
    if current_user.role != "courier":
        return abort(403)
    courier = Courier.query.filter_by(telegram=f"@{current_user.username}").first()
    if not courier:
        return abort(403)
    data = request.get_json() or {}
    ids = data.get("ids", [])
    if not isinstance(ids, list):
        return jsonify(success=False), 400
    orders = (
        Order.query.filter(Order.id.in_(ids))
        .filter_by(courier_id=courier.id, status="Подготовлен к доставке")
        .all()
    )
    for o in orders:
        o.status = "Выдано на доставку"
    db.session.commit()
    return jsonify(success=True)


@app.route("/courier/accept_all", methods=["POST"])
@login_required
def accept_all_orders():
    if current_user.role != "courier":
        return abort(403)
    courier = Courier.query.filter_by(telegram=f"@{current_user.username}").first()
    if not courier:
        return abort(403)
    orders = Order.query.filter_by(
        courier_id=courier.id, status="Подготовлен к доставке"
    ).all()
    for o in orders:
        o.status = "Выдано на доставку"
    db.session.commit()
    return jsonify(success=True)


@app.route("/courier/delivered/<int:order_id>", methods=["POST"])
@login_required
def courier_delivered(order_id):
    if current_user.role != "courier":
        return abort(403)
    courier = Courier.query.filter_by(telegram=f"@{current_user.username}").first()
    order = Order.query.get_or_404(order_id)
    if not courier or order.courier_id != courier.id:
        return abort(403)
    if order.status != "Выдано на доставку":
        return jsonify(success=False), 400
    order.status = "Доставлен"
    order.delivered_at = date.today()
    db.session.commit()
    return jsonify(success=True)


@app.route("/courier/problem/<int:order_id>", methods=["POST"])
@login_required
def courier_problem(order_id):
    if current_user.role != "courier":
        return abort(403)
    courier = Courier.query.filter_by(telegram=f"@{current_user.username}").first()
    order = Order.query.get_or_404(order_id)
    if not courier or order.courier_id != courier.id:
        return abort(403)
    if order.status != "Выдано на доставку":
        return jsonify(success=False), 400
    data = request.get_json() or {}
    comment = data.get("comment") or ""
    order.status = "Проблема"
    order.problem_comment = comment
    db.session.commit()
    return jsonify(success=True)


@app.route("/users")
@admin_required
def users():
    all_users = User.query.all()
    data = []
    for u in all_users:
        zones = ""
        if u.role == "courier":
            c = Courier.query.filter_by(telegram=f"@{u.username}").first()
            zones = c.zones if c and c.zones else ""
        data.append({"user": u, "zones": zones})
    zones_list = DeliveryZone.query.all()
    return render_template("users.html", users=data, zones=zones_list)


@app.route("/users/create", methods=["POST"])
@admin_required
def create_user():
    username = request.form.get("username") or ""
    password = request.form.get("password") or ""
    role = request.form.get("role", "courier")
    zone = request.form.get("zone") or ""
    if not username or not password:
        flash("Введите имя и пароль", "warning")
        return redirect(url_for("users"))
    if role == "courier" and not zone:
        flash("Выберите зону", "warning")
        return redirect(url_for("users"))
    if User.query.filter_by(username=username).first():
        flash("Такой пользователь уже существует", "warning")
        return redirect(url_for("users"))
    user = User(
        username=username,
        password_hash=generate_password_hash(password),
        role=role,
    )
    db.session.add(user)
    if role == "courier":
        courier = Courier(name=username, telegram=f"@{username}", zones=zone)
        db.session.add(courier)
    db.session.commit()
    flash("Пользователь создан", "success")
    return redirect(url_for("users"))


@app.route("/users/<int:user_id>/edit", methods=["POST"])
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    password = request.form.get("password") or ""
    role = request.form.get("role", user.role)
    zone = request.form.get("zone") or ""
    if role == "courier" and not zone:
        flash("Выберите зону", "warning")
        return redirect(url_for("users"))
    if password:
        user.password_hash = generate_password_hash(password)
    user.role = role
    courier = Courier.query.filter_by(telegram=f"@{user.username}").first()
    if role == "courier":
        if not courier:
            courier = Courier(name=user.username, telegram=f"@{user.username}")
            db.session.add(courier)
        courier.zones = zone
    elif courier:
        Order.query.filter_by(courier_id=courier.id).update({"courier_id": None})
        db.session.delete(courier)
    db.session.commit()
    flash("Пользователь обновлён", "success")
    return redirect(url_for("users"))


@app.route("/users/<int:user_id>/delete", methods=["POST"])
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.role == "admin":
        flash("Нельзя удалить администратора", "warning")
        return redirect(url_for("users"))
    courier = Courier.query.filter_by(telegram=f"@{user.username}").first()
    if courier:
        Order.query.filter_by(courier_id=courier.id).update({"courier_id": None})
        db.session.delete(courier)
    db.session.delete(user)
    db.session.commit()
    flash("Пользователь удалён", "success")
    return redirect(url_for("users"))


def _history_query(period):
    q = Order.query.filter_by(status="Доставлен")
    today = date.today()
    if period == "today":
        start = today
    elif period == "7":
        start = today - timedelta(days=7)
    elif period == "month":
        start = today - timedelta(days=30)
    else:
        start = None
    if start:
        q = q.filter(Order.delivered_at >= start)
    return q


@app.route("/history")
@admin_required
def history():
    period = request.args.get("period", "today")
    orders = _history_query(period).all()
    return render_template("history.html", orders=orders, period=period)


@app.route("/history/export")
@admin_required
def export_history():
    period = request.args.get("period", "today")
    orders = _history_query(period).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["№ заказа", "Имя", "Телефон", "Адрес", "Дата", "Зона"])
    for o in orders:
        dt = o.delivered_at.isoformat() if o.delivered_at else ""
        ws.append([o.order_number, o.client_name, o.phone, o.address, dt, o.zone or ""])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"dostavki_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/reports")
@admin_required
def reports():
    batches = [
        b[0] for b in db.session.query(Order.import_batch).distinct().all() if b[0]
    ]
    batches.sort()
    return render_template("reports.html", batches=batches)


@app.route("/reports/export/<string:rtype>")
@admin_required
def export_report(rtype):
    start = request.args.get("start")
    end = request.args.get("end")
    q = Order.query
    if start:
        try:
            start_d = datetime.fromisoformat(start).date()
            q = q.filter((Order.delivered_at == None) | (Order.delivered_at >= start_d))
        except ValueError:
            pass
    if end:
        try:
            end_d = datetime.fromisoformat(end).date()
            q = q.filter((Order.delivered_at == None) | (Order.delivered_at <= end_d))
        except ValueError:
            pass
    if rtype == "delivered":
        q = q.filter_by(status="Доставлен")
    elif rtype == "problem":
        q = q.filter_by(status="Проблема")
    elif rtype != "all":
        abort(400)
    orders = q.all()
    data = []
    for o in orders:
        row = {
            "№ заказа": o.order_number,
            "Имя": o.client_name,
            "Телефон": o.phone,
            "Адрес": o.address,
            "Дата": o.delivered_at.isoformat() if o.delivered_at else "",
            "Зона": o.zone or "",
            "Статус": o.status,
        }
        if rtype in {"problem", "all"}:
            row["Комментарий"] = o.problem_comment or ""
        data.append(row)
    df = pd.DataFrame(data)
    buf = BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    filename = f"report_{rtype}_{date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _orders_to_excel(orders, include_comment=False):
    data = []
    for o in orders:
        row = {
            "№ заказа": o.order_number,
            "Имя": o.client_name,
            "Телефон": o.phone,
            "Адрес": o.address,
            "Дата": o.delivered_at.isoformat() if o.delivered_at else "",
            "Зона": o.zone or "",
        }
        if include_comment:
            row["Комментарий"] = o.problem_comment or ""
        data.append(row)
    df = pd.DataFrame(data)
    buf = BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)
    return buf


@app.route("/download/delivered")
@admin_required
def download_delivered():
    batch = request.args.get("batch")
    if not batch:
        abort(400)
    orders = Order.query.filter_by(import_batch=batch, status="Доставлен").all()
    buf = _orders_to_excel(orders)
    filename = f"delivered_{batch}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/download/problem")
@admin_required
def download_problem():
    batch = request.args.get("batch")
    if not batch:
        abort(400)
    orders = Order.query.filter_by(import_batch=batch, status="Проблема").all()
    buf = _orders_to_excel(orders, include_comment=True)
    filename = f"problem_{batch}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def read_file_rows(path):
    if path.lower().endswith(".csv"):
        with open(path, newline="", encoding="utf-8-sig") as f:
            return [row for row in csv.reader(f)]
    wb = openpyxl.load_workbook(path, read_only=True)
    sheet = wb.active
    rows = []
    for r in sheet.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in r])
    return rows


def _normalize_header(header: str):
    """Collapse various header names to canonical Order field names.

    Unknown headers return ``None`` so they can be ignored during import.
    """
    header = header.strip().lower()
    mapping = {
        "номер заказа": "order_number",
        "номер": "order_number",
        "order number": "order_number",
        "заказ": "order_number",
        "client name": "client_name",
        "имя клиента": "client_name",
        "клиент": "client_name",
        "имя": "client_name",
        "phone": "phone",
        "телефон": "phone",
        "тел": "phone",
        "address": "address",
        "адрес": "address",
        "zone": "zone",
        "зона": "zone",
        "comment": "comment",
        "комментарий": "comment",
        "note": "note",
    }
    return mapping.get(header)


def run_import(job_id, path, batch_name, col_map=None, header=True, clear_old=False):
    """Background import job that processes uploaded files."""
    with app.app_context():
        job = db.session.get(ImportJob, job_id)
        try:
            if clear_old:
                Order.query.delete()
                ImportBatch.query.delete()
                db.session.commit()
                if db.session.bind.dialect.name == "postgresql":
                    inspector = inspect(db.engine)
                    seq_name = None
                    for col in inspector.get_columns(Order.__tablename__):
                        if col["name"] == "id":
                            default = col.get("default") or ""
                            m = re.search(r"nextval\('(.+?)'::regclass\)", default)
                            if m:
                                seq_name = m.group(1)
                            break
                    if seq_name:
                        db.session.execute(
                            text(f"ALTER SEQUENCE {seq_name} RESTART WITH 1")
                        )
                        db.session.commit()
            batch = ImportBatch(name=batch_name)
            db.session.add(batch)
            db.session.commit()

            rows = read_file_rows(path)
            if col_map is None:
                header = True
                header_row = (
                    [h.strip() if isinstance(h, str) else "" for h in rows[0]]
                    if rows
                    else []
                )
                optional_pattern = re.compile(
                    r"\(optional\)|\[optional\]|optional", re.I
                )
                col_map = {}
                for idx, col_name in enumerate(header_row):
                    if not col_name:
                        continue
                    cleaned = optional_pattern.sub("", col_name).strip()
                    canonical = _normalize_header(cleaned.lower())
                    if not canonical:
                        continue
                    col_map[idx] = canonical
                data_rows = rows[1:]
            else:
                data_rows = rows[1:] if header else rows
            job.total_rows = len(data_rows)
            db.session.commit()
            socketio.emit(
                "import_progress",
                {
                    "processed": job.processed_rows,
                    "total": job.total_rows,
                    "status": job.status,
                },
            )

            imported = 0
            for row_num, row in enumerate(data_rows, start=2 if header else 1):
                if all((not c or str(c).strip() == "") for c in row):
                    continue
                try:
                    data = {}
                    for idx, field in col_map.items():
                        value = row[idx] if idx < len(row) else None
                        value = value.strip() if isinstance(value, str) else value
                        if value == "":
                            value = None
                        data[field] = value
                    if not data.get("order_number"):
                        data["order_number"] = f"AUTO-{int(time.time())}"
                    if not data.get("client_name"):
                        data["client_name"] = "Неизвестный клиент"
                    if "address" in data and data["address"]:
                        lat, lon = geocode_address(data["address"])
                        data["latitude"] = lat
                        data["longitude"] = lon
                        data["zone"] = detect_zone(lat, lon) if lat and lon else None
                    order = Order(
                        **data,
                        import_batch=batch_name,
                        import_batch_id=batch.id,
                        import_id=job_id,
                        local_order_number=imported + 1,
                    )
                    db.session.add(order)
                    if order.zone:
                        courier = assign_courier_for_zone(order.zone)
                        if courier:
                            order.courier = courier
                    imported += 1
                    job.processed_rows = imported
                    if imported % 10 == 0:
                        db.session.commit()
                        socketio.emit(
                            "import_progress",
                            {
                                "processed": job.processed_rows,
                                "total": job.total_rows,
                                "status": job.status,
                            },
                        )
                except Exception as exc:
                    app.logger.error("Error importing row %s: %s", row_num, exc)
                    job_errors[job_id].append(f"Строка {row_num}: {exc}")
            job.status = "finished"
        except Exception:
            app.logger.exception("Import job failed")
            job.status = "failed"
        finally:
            job.finished_at = datetime.utcnow()
            db.session.commit()
            socketio.emit(
                "import_progress",
                {
                    "processed": job.processed_rows,
                    "total": job.total_rows,
                    "status": job.status,
                },
            )
            try:
                os.remove(path)
            except Exception:
                pass


@app.route("/api/import/start", methods=["POST"])
@login_required
def api_import_start():
    file = request.files.get("file")
    if not file or file.filename == "":
        return jsonify({"error": "no file"}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in [".csv", ".xlsx"]:
        return jsonify({"error": "bad format"}), 400
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
    uid = str(uuid.uuid4()) + ext
    path = os.path.join(app.config["UPLOAD_FOLDER"], uid)
    file.save(path)
    batch_name = os.path.splitext(file.filename)[0] or os.path.splitext(uid)[0]
    clear_old = bool(request.form.get("clear"))
    job = ImportJob(filename=uid)
    db.session.add(job)
    db.session.commit()
    threading.Thread(
        target=run_import,
        args=(job.id, path, batch_name, None, True, clear_old),
    ).start()
    return jsonify({"job_id": str(job.id)})


@app.route("/api/import/active")
@login_required
def api_import_active():
    job = (
        ImportJob.query.filter(ImportJob.status == "running")
        .order_by(ImportJob.started_at.desc())
        .first()
    )
    if not job:
        return "", 204
    return jsonify(
        {
            "job_id": str(job.id),
            "processed": job.processed_rows,
            "total": job.total_rows,
            "status": job.status,
        }
    )


@app.route("/import/progress/<job_id>")
@login_required
def import_progress(job_id):
    job = ImportJob.query.get_or_404(uuid.UUID(job_id))
    return render_template("progress.html", job_id=job_id, filename=job.filename)


@app.route("/import/status/<job_id>")
@login_required
def import_status(job_id):
    job = ImportJob.query.get_or_404(uuid.UUID(job_id))
    return jsonify(
        {
            "processed": job.processed_rows,
            "total_rows": job.total_rows or 0,
            "status": job.status,
            "errors": job_errors.get(job_id),
        }
    )


@app.route("/import/result/<job_id>")
@login_required
def import_result(job_id):
    job = ImportJob.query.get_or_404(uuid.UUID(job_id))
    return render_template("import_result.html", count=job.processed_rows)


@app.route("/import/upload", methods=["GET", "POST"])
@admin_required
def import_upload():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or file.filename == "":
            flash("Выберите файл", "warning")
            return redirect(url_for("import_upload"))
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in [".csv", ".xlsx"]:
            flash("Недопустимый формат файла", "warning")
            return redirect(url_for("import_upload"))
        uid = str(uuid.uuid4()) + ext
        upload_dir = app.config["UPLOAD_FOLDER"]
        os.makedirs(upload_dir, exist_ok=True)
        path = os.path.join(upload_dir, uid)
        file.save(path)
        job = ImportJob(filename=uid)
        db.session.add(job)
        db.session.commit()
        return redirect(url_for("import_mapping", job_id=job.id))
    return render_template("import_upload.html")


@app.route("/import/mapping/<job_id>")
@admin_required
def import_mapping(job_id):
    job = ImportJob.query.get_or_404(uuid.UUID(job_id))
    path = os.path.join(app.config["UPLOAD_FOLDER"], job.filename)
    rows = read_file_rows(path)
    preview = rows[:5]
    column_count = max(len(r) for r in preview) if preview else 0
    return render_template(
        "import_mapping.html", job_id=job.id, preview=preview, column_count=column_count
    )


@app.route("/import/finish/<job_id>", methods=["POST"])
@admin_required
def import_finish(job_id):
    job = ImportJob.query.get_or_404(uuid.UUID(job_id))
    path = os.path.join(app.config["UPLOAD_FOLDER"], job.filename)
    batch_name = os.path.splitext(job.filename)[0]
    header = bool(request.form.get("header"))
    clear_old = bool(request.form.get("clear"))
    col_map = {}
    for key, value in request.form.items():
        if key.startswith("map_") and value:
            col_map[int(key.split("_")[1])] = value
    threading.Thread(
        target=run_import,
        args=(job.id, path, batch_name, col_map, header, clear_old),
    ).start()
    return "OK"


@app.route("/stats")
@admin_required
def stats():
    zones = DeliveryZone.query.all()
    couriers = Courier.query.all()
    return render_template("stats.html", zones=zones, couriers=couriers)


@app.route("/stats/data")
@admin_required
def stats_data():
    start = request.args.get("start")
    end = request.args.get("end")

    base_q = Order.query
    if start:
        try:
            start_d = datetime.fromisoformat(start).date()
            base_q = base_q.filter(
                (Order.delivered_at == None) | (Order.delivered_at >= start_d)
            )
        except ValueError:
            pass
    if end:
        try:
            end_d = datetime.fromisoformat(end).date()
            base_q = base_q.filter(
                (Order.delivered_at == None) | (Order.delivered_at <= end_d)
            )
        except ValueError:
            pass

    resp = {
        "prepared": base_q.filter(Order.status == "Подготовлен к доставке").count(),
        "out_for_delivery": base_q.filter(Order.status == "Выдано на доставку").count(),
        "delivered": base_q.filter(Order.status == "Доставлен").count(),
        "problem": base_q.filter(Order.status == "Проблема").count(),
    }
    return jsonify(resp)


if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1 and sys.argv[1] in ["db", "migrate", "upgrade", "downgrade"]:
        # \u041a\u043e\u043c\u0430\u043d\u0434\u0430 \u043c\u0438\u0433\u0440\u0430\u0446\u0438\u0438 \u2014 \u043d\u0435 \u0437\u0430\u043f\u0443\u0441\u043a\u0430\u0435\u043c \u0441\u0435\u0440\u0432\u0435\u0440
        pass
    else:
        socketio.run(app, host="0.0.0.0", port=5000)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
